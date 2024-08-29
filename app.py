from flask import Flask, render_template, request, redirect, url_for, session
from werkzeug.security import generate_password_hash, check_password_hash
import mysql.connector
from mysql.connector import connect, Error
import os
from openpyxl import Workbook, load_workbook
from datetime import datetime

app = Flask(__name__)
app.secret_key = os.urandom(24)


db_config = {
    'host': 'localhost',
    'user': 'root',
    'password': 'root',
    'database': 'admission_portal'
}

def create_connection():
    
    try:
        connection = connect(
            host="localhost",
            user="root",
            password="root",
            database="admission_portal"
        )
        print("Database connection successful")
        return connection
    except Error as e:
        print(f"Error connecting to database: '{e}'")
        return None

def init_db():
    connection = create_connection()
    if connection:
        try:
            cursor = connection.cursor()

            cursor.execute('''
                CREATE TABLE IF NOT EXISTS users (
                    email VARCHAR(255) PRIMARY KEY,
                    username VARCHAR(255) NOT NULL UNIQUE,
                    phone VARCHAR(20),
                    password VARCHAR(255) NOT NULL,
                    confirmpassword VARCHAR(255) NOT NULL
                )
            ''')

            connection.commit()
            print("Database initialized successfully")
        except Error as e:
            print(f"Error: '{e}'")
        finally:
            cursor.close()
            connection.close()

init_db()
    

@app.route('/')
def index():
    return render_template('index.html')

    

# @app.route('/create_account', methods=['GET', 'POST'])
# def create_account():
#     if request.method == 'POST':
#         email = request.form['email']
#         username = request.form['username']
#         phone = request.form['phone']
#         password = generate_password_hash(request.form['password'])
        
        
#         connection = create_connection()
#         if connection:
#             try:
#                 cursor = connection.cursor()
#                 cursor.execute("INSERT INTO admission_portal.users VALUES (%s, %s, %s, %s,%s)",
#                                (email, username, phone, password,confirmpassword))
#                 connection.commit()
#                 print(f"User created: {username}, {email}")
#                 return redirect(url_for('login'))
#             except Error as e:
#                 print(f"Error: '{e}'")
#                 return "An error occurred while creating the account."
#             finally:
#                 cursor.close()
#                 connection.close()

  
#     return render_template('create_account.html')
 
def create_connection():
    try:
        connection = connect(
            host="localhost",        # Update with your database host
            user="yourusername",     # Update with your database username
            password="yourpassword", # Update with your database password
            database="admission_portal" # Update with your database name
        )
        print("Database connection successful")
        return connection
    except Error as e:
        print(f"Error connecting to database: '{e}'")
        return None

@app.route('/create_account', methods=['GET', 'POST'])
def create_account():
    if request.method == 'POST':
        email = request.form['email']
        username = request.form['username']
        phone = request.form['phone']
        password = request.form['password']
        confirmpassword = request.form['confirmpassword']

        if password != confirmpassword:
            print("Passwords do not match")
            return "Passwords do not match. Please try again."

        hashed_password = generate_password_hash(password)

        connection = create_connection()
        if connection:
            try:
                cursor = connection.cursor()

                # Specify column names explicitly, without confirmpassword
                cursor.execute("INSERT INTO users (email, username, phone, password) VALUES (%s, %s, %s, %s)",
                               (email, username, phone, hashed_password))
                
                connection.commit()
                print(f"User created: {username}, {email}")
                return redirect(url_for('login'))
            except Error as e:
                print(f"Error inserting user into database: '{e}'")
                return "An error occurred while creating the account."
            finally:
                cursor.close()
                connection.close()
                print("Database connection closed")
        else:
            print("Failed to connect to the database.")
            return "Database connection error."
  
    return render_template('create_account.html')

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        
        connection = create_connection()
        if connection:
            try:
                cursor = connection.cursor(dictionary=True)
                cursor.execute("SELECT * FROM users WHERE username = %s OR email = %s", (username, username))
                user = cursor.fetchone()

                if user and check_password_hash(user['password'], password):
                    session['user_id'] = user['id']
                    print(f"Login successful for user: {username}")
                    return redirect(url_for('submit_admission'))
                else:
                    print(f"Invalid login attempt for user: {username}")
                    return "Invalid username or password"
            except Error as e:
                print(f"Error: '{e}'")
                return "An error occurred during login"
            finally:
                cursor.close()
                connection.close()

    return render_template('login.html')

@app.route('/submit_admission', methods=['GET', 'POST'])
def submit_admission():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    if request.method == 'POST':
        user_id = session['user_id']
        name = request.form['name']
        contact_number = request.form['contact_number']
        father_name = request.form['father_name']
        mother_name = request.form['mother_name']
        address = request.form['address']
        id_proof = request.files['id_proof'].filename
        marksheet = request.files['marksheet'].filename
        fees_paid = float(request.form['fees_paid'])
        payment_date = request.form['payment_date']
        total_amount = float(request.form['total_amount'])
        balance_amount = float(request.form['balance_amount'])
        due_date = request.form['due_date']
        parent_concat = request.form['parent_cotact']


        connection = create_connection()
        if connection:
            try:
                cursor = connection.cursor()
                cursor.execute('''
                    INSERT INTO admissions
                    (user_id, name, contact_number, father_name, mother_name, address,
                    id_proof, marksheet, fees_paid, payment_date, total_amount,
                    balance_amount, due_date, parent_contact)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                ''', (user_id, name, contact_number, father_name, mother_name, address,
                      id_proof, marksheet, fees_paid, payment_date, total_amount,
                      balance_amount, due_date, parent_concat))
                connection.commit()
                print(f"Admission submitted for user: {user_id}")



                excel_file = 'admissions.xlsx'
                if not os.path.exists(excel_file):
                    wb = Workbook()
                    ws = wb.active
                    ws.append(['User ID', 'Name', 'Contact number', 'Father\'s Name', 'Mother\'s Name',
                    'Address', 'ID Proof', 'Marksheet', 'Fees Paid', 'Payment Date',
                    'Total AMount', 'Balance Amount', 'Due Date', 'Parent Contact', 'Admission Date'])
                else:
                    wb = load_workbook(excel_file)
                    ws = wb.active

                ws.append([user_id, name, contact_number, father_name, mother_name, address,
                           id_proof, marksheet, fees_paid, payment_date, total_amount,
                           balance_amount, due_date, parent_concat, datetime.now().strftime('%Y-%m-%d %H:%M:%S')])

                wb.save(excel_file)
                print(f"Admission details aded to Excel file: {excel_file}")

                return "Admission form submitted successfully and details added to Excel sheet"
            except mysql.connector.Error as e:
                print(f"MySQl Error: {e}")
                return f"An error occurred while submitting the admisssion form: {str(e)}"
            except Exception as e:
                print(f"Unexpected error: {e}")
                return f"An unexpected error occurred: {str(e)}"
            finally:
                cursor.close()
                connection.close()
        return render_template('submit_admission.html')

             

if __name__ == '__main__':
    app.run(debug=True)